"""
PACING DATA FLOW: This Python script automates the bi-weekly manual updates to excel pacing documents

This flow requires a bearer token as documented in the Basis API Authentication Flow
  Example script activation: python3 PacingDataFlow.py YOUR_BEARER_TOKEN

Written by Logan Johnson for Heinrich Marketing
V1.0 -- Last tested and updated April 1, 2025

TRANSLATING BUSINESS NEEDS TO CODE  

- Problem: Client manually updated 15+ Excel cells per line item 2x/week  
- Solution: Automated API-to-Excel flow with:  
  • Built-in error handling (vs. manual copy/paste errors)  
  • Daily pacing alerts (previously calculated ad-hoc)  
  
- Next Phase: Database integration planned (this script eased transition)  
"""


import requests
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook

CLIENTS = [
    ("Example Client", "mock-campaign-id", "/mock/path/pacing.xlsx"),
    # Additional clients can be added here
]

BASE_API_URL = "https://api.basis.net/v1"

# Define Basis API Functions

# Function to get all line items for an advertiser

def get_line_items(bearer_token, campaign_id):
    url = f"{BASE_API_URL}/campaigns/{campaign_id}/line_items"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {bearer_token}"
    }
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error getting line items for campaign {campaign_id}: {e}")
        return None

# Function to get stats associated with a line item

def get_line_item_stats(bearer_token, campaign_id, line_item_id):
    url = f"{BASE_API_URL}/stats/line_item?campaign_id={campaign_id}"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {bearer_token}"
    }
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        for item in data.get('data', []):
            if item.get('line_item_id') == line_item_id:
                return item
        return None
    except requests.exceptions.RequestException as e:
        print(f"Error getting stats for line item {line_item_id}: {e}")
        return None

# Function to extract and format relevant stats from the Basis API Response

def extract_relevant_stats(stats_data):
    if not stats_data:
        return None
        
    delivery = stats_data.get('delivery_metrics', {})
    performance = stats_data.get('performance_metrics', {})
    
    # Calculate viewability
    viewable = delivery.get('delivered_viewable_impressions', 0)
    measurable = delivery.get('delivered_measurable_impressions', 1)  # Avoid division by zero
    viewability_pct = (viewable / measurable) * 100 if measurable else 0
    
    return {
        'date_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'data_through_date': stats_data.get('data_through_date'),
        'impressions': delivery.get('delivered_impressions'),
        'clicks': delivery.get('delivered_clicks'),
        'viewability': viewability_pct,
        'pacing_percent': performance.get('pacing_pct_spend'),
        'spend': delivery.get('total_spend'),
        'auctions_won': delivery.get('auctions_won'),
        'click_through_rate': performance.get('click_through_rate')
    }

# Function to process data for a single client

def process_client(bearer_token, client_name, campaign_id, excel_path):
    print(f"\nProcessing client: {client_name}")
    
    line_items_data = get_line_items(bearer_token, campaign_id)
    if not line_items_data:
        print(f"No line items data for {client_name}")
        return
        
    line_items = line_items_data.get('data', [])
    if not line_items:
        print(f"No line items found for {client_name}")
        return
    
    for line_item in line_items:
        line_item_id = line_item.get('id')
        line_item_name = line_item.get('name', '')
        if not line_item_id:
            continue
            
        print(f"\nProcessing line item: {line_item_name}")
        
        stats_data = get_line_item_stats(bearer_token, campaign_id, line_item_id)
        if not stats_data:
            print(f"No stats data for line item {line_item_name}")
            continue
            
        relevant_stats = extract_relevant_stats(stats_data)
        if not relevant_stats:
            print(f"No relevant stats extracted for line item {line_item_name}")
            continue
        
        print(f"Updating Excel with stats for {line_item_name}")
        update_excel_pacing(excel_path, line_item_name, relevant_stats)

# Define Excel Specific Functions

# Function to find pacing doc cell location based on search term and relative position rule

def find_cell_location(sheet, search_term, rule):
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == search_term:
                if rule == "right":
                    return sheet.cell(row=cell.row, column=cell.column + 1)
                elif rule == "below":
                    return sheet.cell(row=cell.row + 1, column=cell.column)
    print(f"Error: Could not find cell for search term '{search_term}'")
    return None

# Function to update excel doc

def update_excel_pacing(excel_path, line_item_name, stats):
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active  # Using active sheet - adjust if specific sheet needed
        
        # Update general metrics
        pacing_through_cell = find_cell_location(sheet, "Pacing Through", "right")
        pacing_through_cell.value = (datetime.now() - timedelta(days=1)).strftime('%m/%d/%y')

        basis_pacing_cell = find_cell_location(sheet, "Basis Pacing % (7 day avg)", "below")
        basis_pacing_cell.value = stats['pacing_percent']

        impressions_cell = find_cell_location(sheet, "Impressions:", "right")
        impressions_cell.value = stats['impressions']

        clicks_cell = find_cell_location(sheet, "Clicks:", "right")
        clicks_cell.value = stats['clicks']

        viewability_cell = find_cell_location(sheet, "Viewability:", "right")
        viewability_cell.value = stats['viewability']
        
        # Update line-item specific metrics
        line_item_found = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column_letter == 'B' and cell.value and line_item_name in str(cell.value):
                    line_item_found = True
                        
                    spend_cell = find_cell_location(sheet, "Spend to Date", "below")
                    if spend_cell:
                        spend_cell.value = stats['spend']
        
        if not line_item_found:
            print(f"Warning: Line item '{line_item_name}' not found in Excel")
        
        workbook.save(excel_path)
        print(f"Successfully updated {excel_path}")
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")

# Main function to process data for all clients

def main(bearer_token):
    if not bearer_token:
        print("Error: No bearer token provided")
        return
        
    for client in CLIENTS:
        client_name, campaign_id, excel_path = client
        process_client(bearer_token, client_name, campaign_id, excel_path)

# Begin flow control and initialize main function

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python PacingDataFlow.py <bearer_token>")
        sys.exit(1)
        
    bearer_token = sys.argv[1]
    main(bearer_token)